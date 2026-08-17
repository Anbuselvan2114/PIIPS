"""
Excel export for PIIPS.

Generates a workbook with three sheets (Purchase Header, Purchase Line,
Reservation Entry) matching the template in "excel format/excel.xlsx".
Which JSON field fills each Excel column is decided by a user-defined
mapping (see the Mapping menu) stored in SQL Server (tbl_FieldMapping).

- Purchase Header : one row per invoice (header-level fields).
- Purchase Line   : one row per line item (item fields, plus header fields
                    like the invoice number for Document No.).
- Reservation Entry : left empty for now (headers preserved).
"""

import os
import re

import openpyxl

from invoice_schema import build_invoice_json


BASE_DIR = os.path.dirname(os.path.abspath(__file__))

TEMPLATE_PATH = os.path.join(BASE_DIR, "excel format", "excel.xlsx")

# In the template, row 1 is a title, row 2 is blank, row 3 holds the column
# names, so data rows start at row 4.
HEADER_ROW = 3

PER_INVOICE_SHEET = "Purchase Header"
# Sheets with one row per line item.
PER_ITEM_SHEETS = {"Purchase Line", "Reservation Entry"}
EMPTY_SHEETS = set()

# Reservation Entry.Source Type = 39 means "Purchase Line" table in BC.
PURCHASE_LINE_SOURCE_TYPE = "39"

# The three sheets, in order. Used when the template file is unavailable
# (columns then come from the DB, tbl_SheetColumn).
KNOWN_SHEETS = ["Purchase Header", "Purchase Line", "Reservation Entry"]


# ---------------------------------------------------------------------------
# Mapping persistence
# ---------------------------------------------------------------------------

def load_mapping():
    """Active {sheet: {column: json_field}} from the DB. Returns {} if the
    database is unavailable so processing/reads degrade gracefully."""
    try:
        import database
        return database.get_field_mapping()
    except Exception:  # noqa: BLE001 - DB optional at read time
        return {}


def save_mapping(mapping, user_id=None):
    import database
    return database.save_field_mapping(mapping, user_id)


# ---------------------------------------------------------------------------
# Template introspection
# ---------------------------------------------------------------------------

def template_columns():
    """Return {sheet_title: [column names]} read from the template row 3.
    Returns {} if the template file is not available."""
    if not os.path.exists(TEMPLATE_PATH):
        return {}
    cols = {}
    wb = openpyxl.load_workbook(TEMPLATE_PATH, read_only=True)
    for ws in wb.worksheets:
        header = []
        for row in ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW,
                                values_only=True):
            header = [str(c).strip() for c in row if c is not None]
        cols[ws.title] = header
    wb.close()
    return cols


def load_columns():
    """Saved column customizations {sheet: [columns]} from the DB (may be
    partial). Returns {} if the database is unavailable."""
    try:
        import database
        return database.get_sheet_columns()
    except Exception:  # noqa: BLE001 - DB optional at read time
        return {}


def save_columns(columns, user_id=None):
    import database
    return database.save_sheet_columns(columns, user_id)


def sheet_columns():
    """
    Effective columns per sheet: the user's customized list (tbl_SheetColumn)
    when present, otherwise the template's columns. Works even if the
    template file is missing, as long as the DB has the sheet.
    """
    tmpl = template_columns()
    custom = load_columns()

    sheets = list(tmpl.keys())
    for s in KNOWN_SHEETS:
        if s not in sheets:
            sheets.append(s)

    result = {}
    for sheet in sheets:
        if isinstance(custom.get(sheet), list):
            result[sheet] = custom[sheet]
        else:
            result[sheet] = tmpl.get(sheet, [])
    return result


def available_fields():
    """
    JSON field names available for mapping, derived from the invoice schema:
      header -> scalar top-level fields (one value per invoice)
      item   -> per-line-item fields
    """
    sample = build_invoice_json(
        {"Fields": {}, "Items": [{}], "TaxSummary": []},
        "sample.pdf",
    )
    header = [
        k for k, v in sample.items()
        if k not in ("items", "sf_items") and not isinstance(v, (list, dict))
    ]
    item = list(sample["items"][0].keys()) if sample.get("items") else []
    return {"header": header, "item": item}


# ---------------------------------------------------------------------------
# Value resolution
# ---------------------------------------------------------------------------

def _resolve(spec, header, item):
    """
    Resolve one column's value from its mapping spec:
      ""        -> blank
      "=text"   -> literal constant 'text'
      "field"   -> item[field] if present else header[field] else blank
    """
    if not spec:
        return ""
    if spec.startswith("="):
        return spec[1:]
    if item is not None and spec in item:
        return item[spec]
    if spec in header:
        return header[spec]
    return ""


# ---------------------------------------------------------------------------
# Row building (shared by Excel export and DB save)
# ---------------------------------------------------------------------------

def _cell_value(inv, item, sheet, smap, col):
    """Mapped column -> JSON value; unmapped -> template static value."""
    spec = smap.get(col, "")
    if spec:
        return _resolve(spec, inv, item)
    static = (inv.get("_static") or {}).get(sheet, {})
    return static.get(col, "")


# A freight/courier/forwarding line has no Navision item master entry of
# its own (see ocr_engine.py's "charge": True lines) - BC books it as a
# fixed service line instead. Only applies to PART invoices; SERVICE
# invoices have no per-line GST Group at all.
_FREIGHT_NO = "FRIEGHT IN"


def _freight_line_override(sheet, col, inv, item):
    """Forced Purchase Line values for a freight/charge item on a PART
    invoice, or None if this row/column isn't one."""
    if sheet != "Purchase Line" or not (item or {}).get("_charge"):
        return None
    if (inv.get("_invoice_type") or "").strip().upper() != "PART":
        return None

    if col == "No.":
        return _FREIGHT_NO
    if col == "GST Group Type":
        return "Service"
    if col == "GST Group Code":
        rate = item.get("TaxPercentage")
        rate_str = f"{rate:g}%" if isinstance(rate, (int, float)) and rate else ""
        return f"Service {rate_str}".strip() if rate_str else "Service"
    return None


def _link_override(sheet, col, inv, item):
    """
    Enforce the cross-sheet relationship keys (BC join chain):

      Purchase Header.Document Type = Purchase Line.Document Type
      Purchase Header.No.           = Purchase Line.Document No.
      Reservation Entry.Source Type     = 39
      Reservation Entry.Source Subtype  = Purchase Line.Document Type
      Reservation Entry.Source ID       = Purchase Line.Document No.
      Reservation Entry.Source Ref. No. = Purchase Line.Line No.

    Returns the forced value for a link column, or None if `col` is not one.
    """
    doc_type = inv.get("Document Type", "")
    doc_no = inv.get("Document No.", "")
    line_no = (item or {}).get("Line_No", "")

    if sheet == "Purchase Header":
        if col == "Document Type":
            return doc_type
        if col == "No.":
            return doc_no
        if col == "Consignment Note No.":
            # BC's field is too short for the full SPRPUR/YYYY/MM/DD-NNNNN
            # buyer's order token (see buyer_order.py) - strip the constant
            # "SPRPUR/" prefix, keeping only the part that's actually
            # unique per order, e.g. "SPRPUR/2026/07/27-85557" ->
            # "2026/07/27-85557".
            return re.sub(r"^[S5]PRPUR/?", "", inv.get("buyer_order_no", "") or "",
                          flags=re.IGNORECASE)
    elif sheet == "Purchase Line":
        if col == "Document Type":
            return doc_type
        if col == "Document No.":
            return doc_no
        if col == "Line No.":
            return line_no
        freight = _freight_line_override(sheet, col, inv, item)
        if freight is not None:
            return freight
    elif sheet == "Reservation Entry":
        if col == "Source Type":
            return PURCHASE_LINE_SOURCE_TYPE
        if col == "Source Subtype":
            # BC's Source Subtype is the numeric Document Type option value,
            # not the text ("Order"/etc.) Purchase Header/Line carry - every
            # purchase invoice PIIPS creates is an Order, whose BC option
            # value is 1, so this is a fixed constant like Source Type.
            return "1"
        if col == "Source ID":
            return doc_no
        if col == "Source Ref. No.":
            return line_no
        # One reservation entry per unit -> quantity is always 1.
        if col in ("Quantity", "Quantity (Base)"):
            return "1"
    return None


def _row(inv, item, sheet, smap, cols):
    """Build one row dict: relationship keys win, else mapped, else static."""
    row = {}
    for c in cols:
        ov = _link_override(sheet, c, inv, item)
        row[c] = ov if ov is not None else _cell_value(inv, item, sheet, smap, c)
    return row


def build_rows(invoices, mapping=None):
    """
    Build the rows for every sheet from the invoices + mapping + static
    values. Returns {sheet: {"columns": [...], "rows": [ {col: value} ] }}.
    Purchase Header -> one row per invoice; Purchase Line -> one per item;
    Reservation Entry -> one row per sf_item (from Service First).
    """
    mapping = mapping or load_mapping()
    columns = sheet_columns()
    result = {}

    for sheet, cols in columns.items():
        smap = mapping.get(sheet, {})
        rows = []

        if sheet == "Reservation Entry":
            # One row per reservation (sf_item) fetched from Service First.
            for inv in invoices:
                for sf in inv.get("sf_items", []):
                    rows.append(_row(inv, sf, sheet, smap, cols))
        elif sheet in PER_ITEM_SHEETS:
            # One row per line item.
            for inv in invoices:
                for item in inv.get("items", []):
                    rows.append(_row(inv, item, sheet, smap, cols))
        else:
            # One row per invoice (Purchase Header).
            for inv in invoices:
                rows.append(_row(inv, None, sheet, smap, cols))

        result[sheet] = {"columns": cols, "rows": rows}

    return result


def build_rows_grouped(invoices, mapping=None):
    """
    Like build_rows, but grouped per invoice so the DB layer can link
    child rows to the inserted Purchase Header Id:

        {
          "columns": {sheet: [cols]},
          "groups": [ {"header": {..}, "lines": [..], "reservations": [..]} ]
        }
    """
    mapping = mapping or load_mapping()
    columns = sheet_columns()
    ph_cols = columns.get("Purchase Header", [])
    pl_cols = columns.get("Purchase Line", [])
    re_cols = columns.get("Reservation Entry", [])
    ph_map = mapping.get("Purchase Header", {})
    pl_map = mapping.get("Purchase Line", {})
    re_map = mapping.get("Reservation Entry", {})

    groups = []
    for inv in invoices:
        header = _row(inv, None, "Purchase Header", ph_map, ph_cols)
        lines, reservations = [], []
        for item in inv.get("items", []):
            lines.append(_row(inv, item, "Purchase Line", pl_map, pl_cols))
        # One reservation row per sf_item (from Service First), not per unit.
        for sf in inv.get("sf_items", []):
            reservations.append(_row(inv, sf, "Reservation Entry", re_map, re_cols))
        groups.append({
            "invoice_no": inv.get("invoice_no", ""),
            "po_number_format": inv.get("PO_Number_Format", ""),
            "header": header,
            "lines": lines,
            "reservations": reservations,
        })

    return {
        "columns": {
            "Purchase Header": ph_cols,
            "Purchase Line": pl_cols,
            "Reservation Entry": re_cols,
        },
        "groups": groups,
    }


# ---------------------------------------------------------------------------
# Mandatory-field validation — final gate before an invoice counts as ready.
# Columns are checked on the fully-resolved row (link-override, then mapped
# JSON field, then template static value) built by build_rows_grouped/_row.
# ---------------------------------------------------------------------------

# "No." (header) / "Document No." (line) / "Source ID" (reservation) are
# deliberately NOT in these lists: they're saved blank on purpose (see
# database.save_grouped) and only become real once the invoice is marked
# Loaded — an invoice otherwise ready must not be downgraded to
# DATA MISMATCH just because that future value isn't set yet.
REQUIRED_HEADER_FIELDS = [
    "Pay-to Name", "Pay-to Address", "Ship-to Name", "Ship-to Address",
    "Posting Date", "Payment Terms Code", "Due Date", "Location Code",
    "Vendor Posting Group", "Vendor Invoice No.", "Buy-from Vendor Name",
    "Buy-from Address", "Ship-to Country/Region Code", "Bal. Account Type",
    "Status", "IC Status", "State", "Structure", "GST Vendor Type",
    "GST Order Address State", "Is Employee",
    "InvoiceNo",
]

REQUIRED_LINE_FIELDS = [
    "Line No.", "Type", "Location Code", "Description",
    "Quantity", "Direct Unit Cost", "Line Amount", "TDS Nature of Deduction",
]

# Purchase Line "No." (Nav Item No.) is mandatory only for an "Item" line —
# it must resolve a Nav Item No. from Service First. A "Charge (Item)" line
# (freight/courier — never looked up in SF) has no Nav Item No. at all, and
# its HSN/SAC Code isn't mandatory either: many vendors simply don't print
# one on a freight/courier line, so requiring it would block otherwise-
# complete invoices over data that was never on the PDF to begin with.
_LINE_TYPE_REQUIRED_FIELD = {
    "Item": "No.",
}

# Source Ref. No. (Reservation Entry) is deliberately NOT in this list: it's
# only resolved once the Purchase Line's own Document No./Line No. become
# real at Load time (see the REQUIRED_HEADER_FIELDS comment above) — it can
# be null before then.
REQUIRED_RESERVATION_FIELDS = [
    "Positive", "Item No.", "Location Code", "Quantity (Base)",
    "Reservation Status", "Creation Date", "Source Type", "Source Subtype",
    "Expected Receipt Date", "Serial No.",
    "Qty. per Unit of Measure", "Quantity", "Suppressed Action Msg.",
    "Planning Flexibility", "Qty. to Handle (Base)", "Qty. to Invoice (Base)",
    "Disallow Cancellation", "Correction", "Item Tracking",
    "Untracked Surplus",
]


def _is_blank(value):
    return value is None or (isinstance(value, str) and not value.strip())


# ---------------------------------------------------------------------------
# Field source (Dashboard "Fields" drill-down popup: where did this value
# come from — Service First / a business Template / the PDF itself).
# ---------------------------------------------------------------------------

# JSON field names Service First (_HEADER_FROM_SF / _enrich_items in
# service_api.py) populates directly.
_SF_HEADER_FIELDS = {
    "Nav_VendorCode", "Pay_to_Vendor_No", "Location_Code", "Location_State_Code",
    "GST_Order_Address_State", "Buy_from_Vendor_No", "Currency_Code",
    "Gen_Bus_Posting_Group", "PaymentTermsName",
}
_SF_LINE_FIELDS = {
    "Nav_Item_No", "ProductNo", "PartSpecification", "HSN_Type",
    "HSN_Percentage_Description", "TaxPercentage", "GST_%", "Nav_Part_Description",
}

# JSON field names computed by processor.py's own business logic rather than
# being a plain passthrough of any single source (payment_terms_code/
# Due_Date combine PaymentTermsName (SF) with a PDF/default fallback through
# resolve_payment_terms()/add_days_to_date() - neither SF nor PDF alone
# determines the final value, the app's rule does).
_SYSTEM_COMPUTED_FIELDS = {"payment_terms_code", "Due_Date"}

# Columns whose value is a computed relationship key (see _link_override)
# rather than a mapped JSON field or template static value — classified by
# what actually determines them, ignoring any configured mapping. "System"
# for ones the app computes outright (a sequence number, a fixed constant,
# BC relationship keys); "Template" only where a template's own static
# value is genuinely the primary driver (with a hardcoded fallback used
# solely when nothing was configured).
_LINK_OVERRIDE_SOURCE = {
    ("Purchase Header", "Document Type"): "Template",
    ("Purchase Header", "No."): "System",
    ("Purchase Header", "Consignment Note No."): "System",
    ("Purchase Line", "Document Type"): "Template",
    ("Purchase Line", "Document No."): "System",
    ("Purchase Line", "Line No."): "System",
    ("Reservation Entry", "Source Type"): "System",
    ("Reservation Entry", "Source Subtype"): "System",
    ("Reservation Entry", "Source ID"): "System",
    ("Reservation Entry", "Source Ref. No."): "System",
    ("Reservation Entry", "Quantity"): "System",
    ("Reservation Entry", "Quantity (Base)"): "System",
}


def field_source(sheet, col, mapping=None):
    """'Service First' / 'Template' / 'PDF' / 'System' / 'None' — where a
    Purchase Header/Line/Reservation Entry column's value comes from, for
    the Dashboard's Fields drill-down popup and the Template screen.
    'System' means the app computes it outright (a business rule, a
    sequence number, a BC relationship key) rather than passing through any
    single source untouched. A reservation row only exists at all because
    of a Service First reservation, so any column actually mapped there
    reads straight from that sf_item data (unlike Purchase Line, its
    mapping specs aren't a PDF/SF mix — every mapped Reservation Entry
    column is Service First). An unmapped column falls back to the
    business Template's static value; everything else is extracted straight
    from the PDF. ('None' — nothing populates it at all — is a Template-
    screen-only refinement layered on top of this by the caller, since it
    depends on whether a static value has actually been entered.)
    Sample: field_source('Purchase Header', 'Location Code')"""
    if col == "InvoiceNo":
        return "PDF"

    override_source = _LINK_OVERRIDE_SOURCE.get((sheet, col))
    if override_source:
        return override_source

    mapping = mapping if mapping is not None else load_mapping()
    spec = (mapping.get(sheet) or {}).get(col, "")
    if not spec or spec.startswith("="):
        return "Template"
    if spec in _SYSTEM_COMPUTED_FIELDS:
        return "System"
    if sheet == "Reservation Entry":
        return "Service First"
    sf_fields = _SF_HEADER_FIELDS if sheet == "Purchase Header" else _SF_LINE_FIELDS
    return "Service First" if spec in sf_fields else "PDF"


def required_fields_for_line_type(line_type):
    """REQUIRED_LINE_FIELDS plus the Type-conditional field ("No." for an
    Item line, "HSN/SAC Code" for a Charge (Item) line), for a given line's
    Type value. Used by both the mandatory-field gate and the Dashboard's
    per-invoice Fields drill-down, so they always agree on what's required.
    Sample: required_fields_for_line_type('Item')"""
    extra = _LINE_TYPE_REQUIRED_FIELD.get((line_type or "").strip())
    return REQUIRED_LINE_FIELDS + [extra] if extra else REQUIRED_LINE_FIELDS


def _line_missing_fields(line):
    return {
        f for f in required_fields_for_line_type(line.get("Type"))
        if _is_blank(line.get(f))
    }


def missing_required_fields(header, lines, reservations=None):
    """Names of mandatory Purchase Header/Line/Reservation Entry columns
    that are empty/null for this invoice (the header row, every line item
    row, and every reservation row). An empty list means the invoice is
    complete. `reservations` defaults to none checked — invoices with no
    reservation rows at all (e.g. NON-GRN, which never syncs with Service
    First) aren't penalized for having none."""
    missing = {f for f in REQUIRED_HEADER_FIELDS if _is_blank(header.get(f))}
    for line in lines:
        missing.update(_line_missing_fields(line))
    for res in (reservations or []):
        missing.update(f for f in REQUIRED_RESERVATION_FIELDS if _is_blank(res.get(f)))
    return sorted(missing)


# ---------------------------------------------------------------------------
# Workbook generation
# ---------------------------------------------------------------------------

def _write_workbook(data, out_path):
    """Write {sheet: {columns, rows}} to an xlsx, preserving the template's
    title/header rows (or a fresh workbook if the template is missing)."""

    if os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    else:
        wb = openpyxl.Workbook()
        default = wb.active
        for name in data.keys():
            wb.create_sheet(name)
        wb.remove(default)

    for ws in wb.worksheets:
        info = data.get(ws.title, {"columns": [], "rows": []})
        cols, rows = info["columns"], info["rows"]

        for ci in range(1, (ws.max_column or 0) + 1):
            ws.cell(row=HEADER_ROW, column=ci, value=None)
        for ci, col in enumerate(cols, start=1):
            ws.cell(row=HEADER_ROW, column=ci, value=col)

        for ri, rowd in enumerate(rows, start=HEADER_ROW + 1):
            for ci, col in enumerate(cols, start=1):
                ws.cell(row=ri, column=ci, value=rowd.get(col, ""))

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path


def build_workbook(invoices, mapping, out_path):
    """Write a workbook directly from invoices + mapping (static/relationship)."""
    return _write_workbook(build_rows(invoices, mapping), out_path)


_DOC_NO_SUFFIX_RE = re.compile(r"^(.*?)(\d+)$")


def renumber_batch(sheet_data, start_doc_no=None, start_entry_no=None):
    """
    Override the auto-generated Document No. / Entry No. in an already-built
    {sheet: {columns, rows}} batch payload before export, per a user-chosen
    starting value (Dashboard > Batches > Document No. / Entry No. inputs).
    Mutates `sheet_data` in place; also returns it.

    Document No.: renumbers Purchase Header["No."] sequentially from
    `start_doc_no`, in row order (fetch_batch orders headers by creation).
    Each header's new value is its own PO_Number_Format (from the template
    active when it was processed - see processor.py/save_grouped) plus the
    running sequence, 6-digit zero-padded (e.g. PO_Number_Format="PO-2627-"
    and start_doc_no=10 -> "PO-2627-000010", start_doc_no=20 ->
    "PO-2627-000020" for the next header). A header with no PO_Number_Format
    on record (no template matched it) instead keeps whatever prefix/padding
    its existing "No." already had, or is just the bare sequence number if
    it never had one either. The same new value is propagated onto
    Purchase Line["Document No."] and
    Reservation Entry["Source ID"] for the rows belonging to that header —
    linked by Purchase_Header_ID (a real FK, present on every row
    database.fetch_batch returns), not by matching the old "No."/
    "Document No."/"Source ID" string values: those are frequently blank
    pre-Load (see excel_export.REQUIRED_*_FIELDS) and were never a
    reliable join key to begin with.

    Entry No.: assigns sequential integers starting at `start_entry_no` to
    every Reservation Entry row across the whole batch, in fetch order.

    Sample: renumber_batch({"Purchase Header": {"rows": [{"Id": 1, "No.": "PO-1"}]}, "Purchase Line": {"rows": []}, "Reservation Entry": {"rows": []}}, start_doc_no=10)
    """
    if start_doc_no is not None:
        header_rows = sheet_data.get("Purchase Header", {}).get("rows", [])
        seq = start_doc_no
        id_to_new_no = {}
        for row in header_rows:
            po_fmt = str(row.get("PO_Number_Format") or "").strip()
            if po_fmt:
                new = f"{po_fmt}{seq:06d}"
            else:
                old = str(row.get("No.") or "").strip()
                m = _DOC_NO_SUFFIX_RE.match(old) if old else None
                new = f"{m.group(1)}{seq:0{len(m.group(2))}d}" if m else str(seq)
            header_id = row.get("Id")
            if header_id is not None:
                id_to_new_no[header_id] = new
            row["No."] = new
            seq += 1

        if id_to_new_no:
            for row in sheet_data.get("Purchase Line", {}).get("rows", []):
                new = id_to_new_no.get(row.get("Purchase_Header_ID"))
                if new is not None:
                    row["Document No."] = new
            for row in sheet_data.get("Reservation Entry", {}).get("rows", []):
                new = id_to_new_no.get(row.get("Purchase_Header_ID"))
                if new is not None:
                    row["Source ID"] = new

    if start_entry_no is not None:
        seq = start_entry_no
        for row in sheet_data.get("Reservation Entry", {}).get("rows", []):
            row["Entry No."] = seq
            seq += 1

    return sheet_data


def build_workbook_from_sheets(data, out_path):
    """Write a workbook from pre-built {sheet: {columns, rows}} (e.g. rows
    read back from the database for a batch)."""
    return _write_workbook(data, out_path)
