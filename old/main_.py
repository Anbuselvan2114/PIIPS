from datetime import datetime, timedelta
import os
import shutil
from datetime import datetime
from openpyxl import load_workbook

from openpyxl import Workbook, load_workbook
from DAO import get_spare_purchase_items_multiple, get_hsn_details
import shutil
import re
import os
import pdfplumber 
import json
from format_detector import detector 

from config import (
    SETTINGS_FILE,
    PATTERN_FILE,
    FORMATS_DIR,
    load_settings,
    INPUT_FOLDER 
)
import shutil
import os
import os
import shutil
import pandas as pd
import json
from datetime import datetime
from collections import defaultdict
import re
from log import add_error,get_errors
from datetime import datetime

from collections import defaultdict
import os

import shutil
from openpyxl import load_workbook
from datetime import datetime
import os
import os
import shutil
from datetime import datetime
from openpyxl import load_workbook

po_counter = 0
settings = load_settings()
header_map = settings.get("ITEMS_HEADER_MAP", {})

print("SETTINGS FILE PATH:", SETTINGS_FILE)

with open(PATTERN_FILE, "r", encoding="utf-8") as f:
    loaded_patterns = json.load(f)

def split_address(address):
    """
    NAV Address = 50 chars
    NAV Address2 = 50 chars
    """

    if not address:
        return "", ""

    address = " ".join(
        str(address)
        .replace("\n", " ")
        .split()
    )

    if len(address) <= 50:
        return address, ""

    split_pos = address.rfind(
        " ",
        0,
        50
    )

    if split_pos <= 0:
        split_pos = 50

    address1 = address[:split_pos].strip()
    address2 = address[split_pos:].strip()

    return (
        address1[:50],
        address2[:50]
    )

def normalize_invoice_date(date_str):
    if not str(date_str).strip():
        return ""

    formats = [
        "%d-%b-%y",      # 28-Apr-26
        "%d-%b-%Y",      # 28-Apr-2026
        "%d-%B-%Y",      # 28-April-2026
        "%d-%m-%Y",      # 10-11-2025
        "%d/%m/%Y",      # 10/11/2025
        "%Y-%m-%d",      # 2025-11-10
        "%Y/%m/%d",
        "%d-%b-%y",      # 8-Nov-25
        "%d-%b-%Y",      # 8-Nov-2025
        "%d-%m-%Y",      # 08-11-2025
        "%d/%m/%Y",      # 08/11/2025
        "%Y-%m-%d"       # 2025-11-08
    ]

    for fmt in formats:
        try:
            return datetime.strptime(
                str(date_str).strip(),
                fmt
            ).strftime("%d-%m-%Y")
        except:
            pass

    return str(date_str).strip()

def safe_number(value):
    if not value:
        return -1
    
    value = str(value).strip()
    
    # remove comma
    value = value.replace(",", "")
    
    # remove trailing dot (example: "1." → "1")
    value = re.sub(r"\.$", "", value)
    
    try:
        return float(value)
    except:
        return -1

def generate_po_number():
    global po_counter
    po_counter += 1
    return str(po_counter)

def read_pdf_text(pdf_path):
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text += page.extract_text() + "\n"
    return text

def build_items_from_body_table(
    Body_Table,
    raw_text,
    header_map,
    po_number
):
    import re

    items_json = []

    hsn_info_global = {
        "Goods": [],
        "Courier": [],
        "Freight": []
    }

    if not Body_Table or len(Body_Table) <= 1:
        return items_json, hsn_info_global

    header = [str(h).strip() for h in Body_Table[0]]

    for row in Body_Table[1:]:

        if not row:
            continue

        if all(not str(c).strip() for c in row):
            continue

        item = {}

        # ------------------------------------------
        # First Description
        # ------------------------------------------
        first_desc = ""

        if len(row) > 1:
            first_desc = clean_description(str(row[1]))

        # ------------------------------------------
        # Find HSN before expanding description
        # ------------------------------------------
        hsn_code = ""

        for i, col in enumerate(header):

            key = header_map.get(col.strip(), col.strip())

            if "hsn" in key.lower():

                if i < len(row):

                    m = re.search(r"\d{4,8}", str(row[i]))

                    if m:
                        hsn_code = m.group(0)
                        break

        # ------------------------------------------
        # Expand description from raw text
        # ------------------------------------------
        full_description = extract_full_description_from_text(
            raw_text=raw_text,
            first_description=first_desc,
            hsn_code=hsn_code
        )

        if not full_description:
            full_description = first_desc

        full_description = clean_description(full_description)

        # Remove tax words
        full_description = re.split(
            r"\b(OUTPUT|IGST|CGST|SGST|CESS)\b",
            full_description,
            flags=re.IGNORECASE
        )[0].strip()

        # ------------------------------------------
        # Build Item JSON
        # ------------------------------------------
        for i, col in enumerate(header):

            key = header_map.get(col.strip(), col.strip())

            if not key:
                continue

            value = ""

            if i < len(row):
                value = str(row[i]).strip()

            # Description
            if key.lower() == "description":
                value = full_description

            # HSN
            elif "hsn" in key.lower():

                if hsn_code:
                    value = hsn_code
                    hsn_info_global["Goods"].append(hsn_code)
                else:
                    value = ""

            # Numeric fields
            elif key.lower() in (
                "rate",
                "amount",
                "quantity",
                "qty",
                "per"
            ):

                m = re.search(r"-?\d[\d,]*\.?\d*", value)

                value = m.group(0).replace(",", "") if m else ""

            item[key] = value

        if not item.get("Description", "").strip():
            continue

        item["Document No."] = po_number
        item["Type"] = "Item"

        items_json.append(item)

    hsn_info_global["Goods"] = sorted(set(hsn_info_global["Goods"]))
    hsn_info_global["Courier"] = sorted(set(hsn_info_global["Courier"]))
    hsn_info_global["Freight"] = sorted(set(hsn_info_global["Freight"]))

    return items_json, hsn_info_global


def clean_description(desc):
    """
    Clean invoice item description without damaging valid model numbers.

    Examples:
    --------------------------------------------------------------------
    TFT 85285200 1 no. 3,200.00 no. 20 INCH MONITOR Freight Outward 9968
        ->
    TFT 20 INCH MONITOR

    PRINTER PARTS & ACCRS. (84439959) CANON GM2070 LOGIC BOARD QM7-6870 SP25117536
        ->
    PRINTER PARTS & ACCRS. (84439959) CANON GM2070 LOGIC BOARD QM7-6870 SP25117536
    --------------------------------------------------------------------
    """

    if not desc:
        return ""

    # Remove PDF artifacts
    desc = re.sub(r'\(cid:\d+\)', ' ', desc)

    # Replace newlines/tabs
    desc = re.sub(r'[\r\n\t]+', ' ', desc)

    # Collapse spaces
    desc = re.sub(r'\s+', ' ', desc).strip()

    # ------------------------------------------------------------------
    # Special handling for monitor descriptions
    # ------------------------------------------------------------------
    if "INCH" in desc.upper():

        # Remove standalone 8-digit HSN (NOT inside parentheses)
        desc = re.sub(r'(?<!\()\b\d{8}\b(?!\))', ' ', desc)

        # Remove quantity + UOM
        desc = re.sub(
            r'\b\d+(?:\.\d+)?\s*(?:no\.?|nos\.?|pcs?|pc|kg|kgs|ltr|ltrs|box|boxes)\b',
            ' ',
            desc,
            flags=re.IGNORECASE
        )

        # Remove prices
        desc = re.sub(r'\b\d[\d,]*\.\d{2}\b', ' ', desc)

        # Remove Freight onwards
        desc = re.sub(
            r'Freight\s+Outward.*',
            '',
            desc,
            flags=re.IGNORECASE
        )

        # Remove trailing SAC code
        desc = re.sub(r'\s+\d{4,8}$', '', desc)

        # Example:
        # TFT 3 20 INCH MONITOR
        desc = re.sub(
            r'(\b[A-Za-z]+\b)\s+\d+\s+(\d+\s+INCH\b)',
            r'\1 \2',
            desc,
            flags=re.IGNORECASE
        )

        # Remove standalone amount fragments before INCH
        desc = re.sub(
            r'\b\d+\b(?=\s+\d+\s+INCH)',
            '',
            desc
        )

        desc = re.sub(r'\s+', ' ', desc).strip()

    return desc
 
 
def extract_full_description_from_text(
        raw_text,
        first_description="",
        hsn_code=None):
    """
    Expands truncated description from invoice text.

    Example:

    PDF

    1 TFT 85285200 1 no. 2800.00 no. 2800.00
      19 INCH MONITOR
    2 VGA CABLE 8544 1 no. 150.00

    Returns

    TFT 19 INCH MONITOR
    """

    if not raw_text:
        return clean_description(first_description)

    first_description = clean_description(first_description)

    if not first_description:
        return ""

    text = raw_text.replace("\r", "\n")
    text = re.sub(r"\(cid:\d+\)", " ", text)

    lines = [
        clean_description(x)
        for x in text.split("\n")
        if x.strip()
    ]

    description = []

    collecting = False

    for line in lines:

        if not collecting:

            if first_description.lower() not in line.lower():
                continue

            collecting = True

            idx = line.lower().find(first_description.lower())

            line = line[idx:]

            # Remove HSN onwards
            if hsn_code:

                line = re.sub(
                    rf"\b{re.escape(str(hsn_code))}\b.*",
                    "",
                    line,
                    flags=re.I
                )

            # Remove quantity/rate columns
            line = re.sub(
                r"\b\d+\s*(NO\.?|NOS\.?|PCS?|PC|KG|KGS|BOX|SET|EA|EACH)\b.*",
                "",
                line,
                flags=re.I
            )

            # Remove price
            line = re.sub(
                r"\d{1,3}(?:,\d{3})+\.\d{2}.*",
                "",
                line
            )

            line = clean_description(line)

            if line:
                description.append(line)

            continue

        # ---------------------------------------------------
        # STOP : next item begins
        # Example:
        # 2 VGA CABLE ...
        # 3 HDD ...
        # ---------------------------------------------------
        if re.match(r"^\d+\s+[A-Za-z]", line):
            break

        # Stop at HSN line
        if hsn_code:

            if line == str(hsn_code):
                break

            if line.startswith(str(hsn_code)):
                break

        # HSN only
        if re.match(r"^\d{4,8}$", line):
            break

        # Qty line
        if re.match(
                r"^\d+(\.\d+)?\s*(NO\.?|NOS\.?|PCS?|PC|KG|KGS|BOX|SET|EA|EACH)\b",
                line,
                re.I):
            break

        # Amount only
        if re.match(r"^[\d,]+\.\d{2}$", line):
            break

        # Totals/Freight
        if re.search(
                r"FREIGHT|COURIER|OUTPUT|CGST|SGST|IGST|CESS|TOTAL|SUB\s*TOTAL|ROUND\s*OFF|GRAND\s*TOTAL",
                line,
                re.I):
            break

        # Ignore lines made only of numbers
        if re.fullmatch(r"[\d\s,\.]+", line):
            continue

        description.append(line)

    if not description:
        return first_description

    desc = clean_description(" ".join(description))

    # Remove duplicate consecutive words
    words = desc.split()

    final = []

    for w in words:

        if not final or final[-1].lower() != w.lower():
            final.append(w)

    return " ".join(final)



def extract_hsn_block(raw_text):
    # Original HSN/SAC block pattern
    pattern1 = re.compile(
        r"(HSN/SAC Taxable[\s\S]*?Tax Amount\s*\(in words\)[^\n]*)",
        re.IGNORECASE
    )
    match1 = pattern1.search(raw_text)
    if match1:
        return match1.group(1).strip()

    # Alternate HSN/SAC pattern (some PDFs use 'in Words')
    pattern2 = re.compile(
        r"(HSN/SAC Taxable[\s\S]*?in words[^\.]*\.)",
        re.IGNORECASE
    )
    match2 = pattern2.search(raw_text)
    if match2:
        return match2.group(1).strip()

    # New simpler tax table format: Tax Rate Taxable Amt. IGST Amt. Total Tax
    pattern3 = re.compile(
        r"(Tax Rate\s+Taxable Amt\.\s+IGST Amt\.\s+Total Tax[\s\S]*?(?:\d+[\d,]*\.\d+\s+){3,})",
        re.IGNORECASE
    )
    match3 = pattern3.search(raw_text)
    if match3:
        return match3.group(1).strip()

    # If nothing matches
    return None
    

def parse_hsn_table(text):
    """
    Parse HSN/SAC tax tables in various formats:
    1. CGST/SGST: HSN Taxable Central Tax Rate Amount State Tax Rate Amount Total
    2. IGST only: HSN Taxable IGST Rate Amount Total
    3. Simple Tax: Tax Rate Taxable Amt. IGST Amt. Total Tax

    Returns structured JSON with:
    HSN, Taxable, IGST_Rate_%, CGST_Rate_%, SGST_Rate_%, IGST_Amount, CGST_Amount, SGST_Amount, Total
    """
    if not text or not text.strip():
        return []

    lines = [l.strip() for l in text.split("\n") if l.strip()]
    if len(lines) < 2:
        return []

    data_rows = []

    # Detect table type from header
    header_line = lines[0].lower()
    is_cgst_sgst = "central" in header_line and "state" in header_line
    is_igst_only = "igst" in header_line and "central" not in header_line
    is_simple_tax = "tax rate" in header_line

    # Skip header lines
    body_lines = lines[2:] if len(lines) > 2 else lines[1:]

    for line in body_lines:
        parts = line.split()
        if not parts:
            continue

        row = {
            "HSN": "",
            "Taxable": "",
            "IGST_Rate_%": "",
            "CGST_Rate_%": "",
            "SGST_Rate_%": "",
            "IGST_Amount": "",
            "CGST_Amount": "",
            "SGST_Amount": "",
            "Total": ""
        }

        # Detect Total row
        is_total_row = parts[0].lower() == "total"

        if is_cgst_sgst and len(parts) >= 7:
            if parts[0]!="Tax":
                row["HSN"] = parts[0] if not is_total_row else "Total"
                row["Taxable"] = parts[1]
                row["CGST_Rate_%"] = parts[2]
                row["CGST_Amount"] = parts[3]
                row["SGST_Rate_%"] = parts[4]
                row["SGST_Amount"] = parts[5]
                row["Total"] = parts[6]

        elif is_igst_only and len(parts) >= 5:
            if parts[0]!="Tax":
                row["HSN"] = parts[0] if not is_total_row else "Total"
                row["Taxable"] = parts[1]
                row["IGST_Rate_%"] = parts[2]
                row["IGST_Amount"] = parts[3]
                row["Total"] = parts[4]

        elif is_simple_tax and len(parts) >= 4:
            # Example: 18% 1,400.00 252.00 252.00
            if parts[0]!="Tax":
                row["HSN"] =""  # Tax Rate goes to HSN column
                row["Taxable"] = parts[1]
                row["IGST_Rate_%"] = parts[0]
                row["IGST_Amount"] = parts[2]
                row["Total"] = parts[3]

        # Normalize Total row
        if is_total_row:
            row["HSN"] = "Total"
            if is_cgst_sgst:
                row["CGST_Amount"] = parts[2] if len(parts) > 2 else ""
                row["SGST_Amount"] = parts[3] if len(parts) > 3 else ""
                row["Total"] = parts[4] if len(parts) > 4 else ""
            elif is_igst_only or is_simple_tax:
                row["IGST_Amount"] = parts[2] if len(parts) > 2 else ""
                row["Total"] = parts[3] if len(parts) > 3 else row["IGST_Amount"]
        
        if parts[0]!="Tax":
            data_rows.append(row)

    return data_rows

 
def assign_tax_from_tax_total(items, tax_total):
    """
    Assign TaxPercentage to items by matching item HSN
    with tax_total HSN.

    Returns updated items.
    """

    if not items or not tax_total:
        return items

    tax_map = {}

    for row in tax_total:

        hsn = str(row.get("HSN", "")).strip()

        if not hsn or hsn.lower() == "total":
            continue

        rate = (
            row.get("IGST_Rate_%")
            or row.get("CGST_Rate_%")
            or row.get("SGST_Rate_%")
            or ""
        ).replace("%", "").strip()

        if rate:
            tax_map[hsn] = rate

    for item in items:

        item_hsn = str(item.get("hsn", "")).strip()

        if item_hsn in tax_map:
            item["TaxPercentage"] = tax_map[item_hsn]

    return items
    
def extract_default_tax_percentage(raw_text):
    """
    Returns the GST percentage found in the invoice.
    Works with:
      - IGST/CGST/SGST rows in item table
      - Tax summary table without HSN
    """

    patterns = [

        # IGST - Sales 18 %
        r'(?:IGST|CGST|SGST).*?(\d+(?:\.\d+)?)\s*%',

        # Taxable Value 18%
        r'Taxable\s+.*?(\d+(?:\.\d+)?)\s*%',

        # 5310.00 18% 955.80
        r'[\d,]+\.\d+\s+(\d+(?:\.\d+)?)%\s+[\d,]+\.\d+',
    ]

    for pattern in patterns:

        m = re.search(
            pattern,
            raw_text,
            flags=re.IGNORECASE | re.DOTALL
        )

        if m:
            return float(m.group(1))

    return 0

def get_freight_items(raw_text, po_number, buyer_order_no, start_line_no=100000):
    """
    Extract Freight/Courier/Transport charges directly from PDF text.

    Args:
        raw_text (str): OCR/PDF text.
        po_number (str): Purchase Order Number.
        start_sl_no (int): Starting serial number.
        start_line_no (int): Starting line number.

    Returns:
        list: Freight/Courier item dictionaries.
    """

    keywords = [
        "freight",
        "freight charges",
        "courier",
        "courier charges",
        "transport",
        "transportation",
        "transportation charges",
        "delivery charges",
        "packing charges",
        "handling charges",
        "cartage"
    ]

    items = []

    sl_no = start_line_no
    line_no = start_line_no

    lines = raw_text.splitlines()

    for line in lines:

        text = " ".join(line.split())

        if not text:
            continue

        lower = text.lower()

        if not any(keyword in lower for keyword in keywords):
            continue

        # Last numeric value in line = amount
        nums = re.findall(r"-?\d[\d,]*\.?\d*", text)

        if not nums:
            continue

        try:
            amount = float(nums[-1].replace(",", ""))
        except Exception:
            continue

        if amount <= 0:
            continue

        # ---------------------------------
        # Extract HSN/SAC
        # Example:
        # Courier Charges - Interstate 9968 120.00
        # -> HSN = 9968
        # ---------------------------------
        hsn = ""

        m = re.search(
            r'(\d{4,8})\s+\d+(?:,\d+)*(?:\.\d+)?\s*$',
            text
        )

        if m:
            hsn = m.group(1)

        items.append({
            "sl_no": line_no,                     
            "Description": text,
            "hsn": hsn,
            "ProductNo":hsn,
            "rate": amount,
            "amount": amount,
            "Document No.": po_number,
            "Type": "Charge (Item)",
            "PurchaseOrderNo": buyer_order_no,
            "PartSpecification": text, 
            "Quantity": 1,
            "Amount": amount,
            "GST_Base_Amount": amount,
            "Line_No": line_no,
            "Nav_Item_No":"FRIEGHT IN",
            "HSN_Type":"Service",
            "HSN_Percentage_Description": f"Service",
        })

        sl_no += 1
        line_no += 10000

    return items

from pathlib import Path

def extract_entity_template(
    pdf_path,
    pdf_root
):

    pdf_path = Path(
        pdf_path
    ).resolve()

    pdf_root = Path(
        pdf_root
    ).resolve()

    try:

        relative = pdf_path.relative_to(
            pdf_root
        )

        parts = relative.parts

        # PT\Services_Chennai\file.pdf

        if len(parts) >= 3:

            entity = parts[0]
            template = parts[1]

            return (
                entity,
                template
            )

    except Exception as e:

        print(
            f"Entity extraction failed: {e}"
        )

    return (
        "Unknown",
        "Default"
    )
def process_pdf(PDF_PATH,session_id=None):
    """
    Process a single PDF file and extract invoice data.
    Returns:
        invoice_data (dict), status (1 = success, 0 = fail)
    """

    po_number = generate_po_number()

    try:

        import re
        import pdfplumber

        Body_Table = []
        formdata = []
        invoice_data = {}

        # =====================================================
        # READ PDF TEXT
        # =====================================================
        raw_text = read_pdf_text(PDF_PATH)

        print(raw_text)
        # =====================================================
        # DETECT FORMAT
        # =====================================================
        # =====================================================
        # DETECT FORMAT
        # =====================================================
        format_name = detector.detect_format(raw_text)

         

        entity = ""
        template = ""
  
        entity, template = extract_entity_template(PDF_PATH,settings["PDF_FOLDER"])

        print("Input Folder   :", PDF_PATH)
        print("Entity         :", entity)
        print("Template       :", template)
        

        if not format_name:

            print(f"⚠ Format not detected for {PDF_PATH}")

            add_error( 
                session_id=session_id,
                file_name=os.path.basename(PDF_PATH),
                entity="",
                invoice_no="",
                reason="PDF format not matched / Unable to detect invoice format"
            ) 

            return {}, 0

        print(
            f"✅ Detected format: "
            f"{format_name}"
        )

        # =====================================================
        # LOAD PATTERNS
        # =====================================================
        pattern_dict = loaded_patterns.get(
            format_name
        )

        if not pattern_dict:

            print(
                f"⚠ No patterns found "
                f"for format {format_name}"
            )

            reason = f"No extraction patterns found for format '{format_name}'. kinldy, contact technical team."
            add_error( 
                session_id=session_id,
                file_name=os.path.basename(PDF_PATH),
                entity="",
                invoice_no="",
                reason=reason
            )

            print(
                "Available formats:",
                list(loaded_patterns.keys())
            )

            return {}, 0

        # =====================================================
        # COMPILE PATTERNS
        # =====================================================
        compiled_patterns = {}

        for key, pattern_list in pattern_dict.items():

            compiled_patterns[key] = []

            if isinstance(pattern_list, list):

                for p in pattern_list:

                    compiled_patterns[key].append(

                        re.compile(
                            p,
                            re.IGNORECASE
                            | re.MULTILINE
                        )
                    )

            else:

                compiled_patterns[key].append(

                    re.compile(
                        pattern_list,
                        re.IGNORECASE
                        | re.MULTILINE
                    )
                )

        # =====================================================
        # EXTRACT TABLES
        # =====================================================
        with pdfplumber.open(PDF_PATH) as pdf:

            for page in pdf.pages:

                page_text = (
                    page.extract_text()
                    or ""
                )

                # =============================================
                # SKIP DUPLICATE PAGE
                # =============================================
                if (
                    "DUPLICATE FOR TRANSPORTER"
                    in page_text.upper()
                ):
                    continue

                tables = (
                    page.extract_tables()
                    or []
                )

                for table in tables:

                    for row in table:

                        cleaned_row = [

                            cell.replace(
                                "\n",
                                " "
                            ).strip()

                            if cell else ""

                            for cell in row
                        ]

                        if not cleaned_row:
                            continue

                        first_cell_text = (
                            cleaned_row[0]
                        )

                        num = safe_number(
                            first_cell_text
                        )

                        is_header = (

                            first_cell_text.strip()

                            in [

                                "Sl No.",
                                "S.N.",
                                "Sr. No.",
                                "Sl",
                                "No."
                            ]
                        )

                        is_serial = (
                            0 < num < 500
                        )

                        # =====================================
                        # HEADER
                        # =====================================
                        if is_header:

                            if not Body_Table:

                                Body_Table.append(
                                    cleaned_row
                                )

                            continue

                        # =====================================
                        # ITEM ROW
                        # =====================================
                        if is_serial:

                            Body_Table.append(
                                cleaned_row
                            )

                        # =====================================
                        # FORM DATA
                        # =====================================
                        for cell in row:

                            if (
                                cell
                                and cell.strip()
                            ):

                                formdata.append(
                                    cell.strip()
                                )

        # =====================================================
        # REMOVE DUPLICATE ROWS
        # =====================================================
        unique_rows = []

        seen = set()

        for row in Body_Table:

            key = tuple(row)

            if key not in seen:

                seen.add(key)

                unique_rows.append(row)

        Body_Table = unique_rows

        # =====================================================
        # FULL TEXT
        # =====================================================
        text = "\n".join(formdata)

        full_text = (
            text
            + "\n"
            + raw_text
        )

        # =====================================================
        # EXTRACT HEADER FIELDS
        # =====================================================
        for key, pattern_list in compiled_patterns.items():

            invoice_data[key] = ""

            for pattern in pattern_list:

                match = pattern.search(
                    full_text
                )

                if match:

                    for group in match.groups():

                        if (
                            group
                            and group.strip()
                        ):

                            invoice_data[key] = (
                                group.strip()
                            )

                            break

                if invoice_data[key]:
                    break

        # =====================================================
        # ITEMS
        # =====================================================
        items_json, hsn_info_global = build_items_from_body_table(
            Body_Table=Body_Table,
            raw_text=raw_text,
            header_map=header_map,
            po_number=po_number
        )

        # Buyer Order No extracted from invoice header
        buyer_order_no = (
            invoice_data.get("buyer_order_no")
            or invoice_data.get("Buyer Order No")
            or ""
        ).strip()

        if buyer_order_no=="Dated":
            invoice_data["buyer_order_no"]=""
            buyer_order_no==""

            
        if invoice_data["buyer_order_no"]=="":

            add_error( 
                session_id=session_id,
                file_name=os.path.basename(PDF_PATH),
                entity=entity,
                invoice_no=invoice_data.get("invoice_no", ""),
                reason="Buyer Order No not found or invalid. Expected value starting with 'SPRPUR'."
            )


            # =====================================
            # MOVE PDF TO PONO_NotValid
            # =====================================
            try:

                pdf_folder = settings["PDF_FOLDER"]

                # Remove last folder (PDF)
                root_folder = os.path.dirname(
                    os.path.normpath(pdf_folder)
                )

                # Create:
                # \\Server\...\Reader\PONO_NotValid
                invalid_folder = os.path.join(
                    root_folder,
                    "PONO_NotValid"
                )

                os.makedirs(
                    invalid_folder,
                    exist_ok=True
                )

                destination = os.path.join(
                    invalid_folder,
                    os.path.basename(PDF_PATH)
                )

                # Overwrite existing file if present
                if os.path.exists(destination):
                    os.remove(destination)

                shutil.move(
                    PDF_PATH,
                    destination
                )

                print(
                    f"📁 Moved to invalid PO folder: {destination}"
                )

            except Exception as move_error:

                print(
                    f"❌ Failed to move PDF: {move_error}"
                )
        # Update Line_No, sl_no and PurchaseOrderNo
        for item in items_json:

            try:
                sl = int(float(item.get("sl_no", 0)))
            except Exception:
                sl = 0

            line_no = sl * 10000

            item["sl_no"] = line_no
            item["Line_No"] = line_no
            item["PurchaseOrderNo"] = buyer_order_no

             # Convert quantity -> Quantity
            try:
                qty = str(item.get("quantity", "0")).replace(",", "").strip()
                item["Quantity"] = int(float(qty))
            except (ValueError, TypeError):
                item["Quantity"] = 0

            # Convert amount -> Amount & GST_Base_Amount
            try:
                amt = str(item.get("amount", "0")).replace(",", "").strip()
                amount = int(float(amt))
            except (ValueError, TypeError):
                amount = 0

            item["Amount"] = amount
            item["GST_Base_Amount"] = amount
               
            # item["HSN_Percentage_Description"] = f"GOODS {gst_percent}%",

        invoice_data["hsn_type"] = hsn_info_global
        invoice_data["items"] = items_json
        
        
 

        # =====================================================
        # ADD FREIGHT / COURIER ITEMS
        # =====================================================
        extra_items = get_freight_items(
            raw_text=raw_text,
            po_number=po_number,
            buyer_order_no=buyer_order_no,
            start_line_no=(len(items_json) + 1) * 10000
        )

        invoice_data["items"].extend(extra_items)

        # =====================================================
        # REMOVE DUPLICATE HSN
        # =====================================================
        for k in hsn_info_global:

            hsn_info_global[k] = list(
                set(hsn_info_global[k])
            )

        invoice_data["hsn_type"] = (
            hsn_info_global
        )

        invoice_data["items"] = (
            items_json
        )

        # =====================================================
        # HSN / TAX SECTION
        # =====================================================
        hsn_block = extract_hsn_block(
            raw_text
        )

        hsn_rows = parse_hsn_table(
            hsn_block
        )

        invoice_data["tax_total"] = [

            row

            for row in hsn_rows

            if "HSN/SAC" not in row
        ]

        # =====================================================
        # ASSIGN ITEM TAX PERCENTAGE
        # =====================================================

        invoice_data["items"] = assign_tax_from_tax_total(
            invoice_data.get("items", []),
            invoice_data.get("tax_total", [])
        )

        default_tax = extract_default_tax_percentage(raw_text)

        for item in invoice_data["items"]:

            # Assign default tax if missing
            if not item.get("TaxPercentage"):
                item["TaxPercentage"] = default_tax

            # Convert TaxPercentage to integer
            try:
                tax = int(float(str(item["TaxPercentage"]).replace("%", "").strip()))
            except (ValueError, TypeError):
                tax = 0

            item["TaxPercentage"] = tax

            # Update HSN Percentage Description
            if tax > 0:
                item["HSN_Percentage_Description"] = f"Service {tax}%"

        invoice_data["invoice_date"] = normalize_invoice_date(
            invoice_data.get("invoice_date", "")
        )
        # # =====================================================
        # # ADDITIONAL CHARGES
        # # =====================================================
        # invoice_data["items"] = (

        #     add_additional_charge_items(

        #         items_json=invoice_data["items"],

        #         invoice_data=invoice_data,

        #         hsn_info_global=hsn_info_global,

        #         po_number=po_number
        #     )
        # )

        
        # =====================================================
        # STORE ENTITY & TEMPLATE
        # =====================================================
        invoice_data["Entity"] = entity
        invoice_data["Template"] = template
        invoice_data["Template_Name"] = format_name
        invoice_data["File_Name"] = os.path.basename(PDF_PATH)

        print("\n====================================")
        print("📦 EXTRACTED ITEMS")
        print("====================================")


        # =====================================================
        # DEBUG
        # =====================================================
        # for itm in invoice_data["items"]:

        #     print(
        #         itm.get(
        #             "Description",
        #             ""
        #         )
        #     )

        # print(invoice_data)
        return invoice_data, 1

    except Exception as e:

        print(
            f"❌ Error processing PDF "
            f"'{PDF_PATH}': {e}"
        )

        return {}, 0

def fetch_api_items(order_numbers, json_data_list): 
    
    print("\n====================================")
    print("🚀 fetch_api_items STARTED")
    print("====================================")

    try:

        

        if not order_numbers:
            print("❌ EMPTY ORDER NUMBERS")
            return {}

        # =====================================================
        # LOAD STATE CODE JSON
        # =====================================================
        with open(
            "statecode.json",
            "r",
            encoding="utf-8"
        ) as f:

            STATE_CODES = json.load(f)

        # print("STATE_CODES")
        # print(STATE_CODES)
        # =====================================================
        # HELPER
        # =====================================================
        def clean(value):

            return str(value or "").strip().lower()

        # =====================================================
        # CLEAN ORDERS
        # =====================================================
        cleaned_orders = list(set(

            clean(x)

            for x in order_numbers

            if clean(x)

        ))

        print(f"📊 UNIQUE ORDER COUNT: {len(cleaned_orders)}")

        # =====================================================
        # SPARE API CALL
        # sf_items[] ONLY
        # =====================================================
        api_response = get_spare_purchase_items_multiple(
            cleaned_orders
        )

        print("====================================")
        print("📥 SPARE API OUTPUT")
        print("====================================")

        for order_no, rows in api_response.items():
            print(f"\nOrder No : {order_no}")
            print(f"Total Rows: {len(rows)}")

            for i, row in enumerate(rows, start=1):
                print(f"\n----- Row {i} -----")

                for key, value in row.items():
                    print(f"{key}: {value}")

                print("-------------------")

        if not api_response:
            return {}

        # =====================================================
        # NORMALIZE SPARE RESPONSE
        # =====================================================
        grouped_data = {}

        if isinstance(api_response, dict):

            if all(
                isinstance(v, list)
                for v in api_response.values()
            ):

                grouped_data = {

                    clean(k): v

                    for k, v in api_response.items()

                }

            else:

                grouped_data = api_response

        # =====================================================
        # BUILD HSN INPUT
        # =====================================================
        hsn_items = []

        for obj in json_data_list:

            data = obj.get("data", {})

            order_no = clean(
                data.get("buyer_order_no")
            )

            for item in data.get("items", []):

                desc = str(
                    item.get("Description", "")
                ).strip()

                if not desc:
                    continue

                hsn_items.append({

                    "PartSpecification":
                        desc,

                    "PurchaseOrderNo":
                        order_no
                })

        # =====================================================
        # REMOVE DUPLICATES
        # =====================================================
        hsn_items = list({

            (
                clean(x["PartSpecification"]),
                clean(x["PurchaseOrderNo"])
            ): x

            for x in hsn_items

        }.values())

        print("📤 HSN API INPUT")
        print(hsn_items)

        # =====================================================
        # CALL HSN API
        # items[] ONLY
        # =====================================================
        hsn_map_raw = get_hsn_details(
            hsn_items
        ) or {}

        print("📥 HSN API OUTPUT")
        print(hsn_map_raw)

        # =====================================================
        # BUILD HSN MAP
        # =====================================================
        hsn_map = {}

        if isinstance(hsn_map_raw, list):

            for row in hsn_map_raw:

                if not isinstance(row, dict):
                    continue

                po = clean(
                    row.get("PurchaseOrderNo")
                )

                part = clean(
                    row.get("PartSpecification")
                )

                if po and part:

                    hsn_map[(po, part)] = row

        elif isinstance(hsn_map_raw, dict):

            for _, row in hsn_map_raw.items():

                if not isinstance(row, dict):
                    continue

                po = clean(
                    row.get("PurchaseOrderNo")
                )

                part = clean(
                    row.get("PartSpecification")
                )

                if po and part:

                    hsn_map[(po, part)] = row

        # =====================================================
        # APPLY DATA
        # =====================================================
        for obj in json_data_list:

            data = obj.get("data", {})

            order_no = clean(
                data.get("buyer_order_no")
            )

            api_rows = grouped_data.get(
                order_no,
                []
            )
          

                 

            # =================================================
            # STATECODE PROCESS
            # =================================================
            buyer_gstin = str(
                data.get("buyer_gstin", "")
            ).strip()

            seller_gstin = str(
                data.get("seller_gstin", "")
            ).strip()

            buyer_state_code = ""
            seller_state_code = ""

            if len(buyer_gstin) >= 2:
                buyer_state_code = buyer_gstin[:2]

            if len(seller_gstin) >= 2:
                seller_state_code = seller_gstin[:2]

            buyer_state_info = STATE_CODES.get(
                buyer_state_code,
                {}
            )

            seller_state_info = STATE_CODES.get(
                seller_state_code,
                {}
            )

            # =================================================
            # HEADER LEVEL
            # =================================================
            data["buyer_state_GST_code"] = buyer_state_code

            data["seller_state_GST_code"] = seller_state_code

            data["seller_state"] = seller_state_info.get(
                "state",
                ""
            )

            data["buyer_state_short"] = buyer_state_info.get(
                "short",
                ""
            )

            data["state"] = seller_state_info.get(
                "short",
                ""
            )

            data["seller_state_short"] = seller_state_info.get(
                "short",
                ""
            )
            # =====================================================
            # APPLY DEFAULTS TO sf_items
            # =====================================================
            today_date = datetime.now().strftime("%Y-%m-%d")

            for sf in api_rows:

                sf["Creation Date"] = today_date

                sf["Expected Receipt Date"] = today_date 

            # =================================================
            # sf_items[] ONLY
            # =================================================
            data["sf_items"] = api_rows

            # =================================================
            # ROOT VALUES FROM sf_items
            # =================================================
            if api_rows:

                first_sf = api_rows[0]
                # if not str(first_sf.get("Nav_VendorCode", "")).strip(): 
                #     error_rows.append({
                #         "FileName": os.path.splitext(
                #             obj.get("file", "")
                #         )[0] + ".pdf",
                #         "Entity": obj.get("entity"),
                #         "InvoiceNo": data.get("invoice_no", ""),
                #         "Reason":
                #             "Navision Vendor Code not mentioned in Vendor Master in Service First"
                #     })

                data["Nav_VendorCode"] = first_sf.get(
                    "Nav_VendorCode",
                    ""
                )

                data["Pay_to_Vendor_No"] = first_sf.get(
                    "Nav_VendorCode",
                    ""
                )

                data["Location_Code"] = first_sf.get(
                    "Location_Code",
                    ""
                )

                data["Location_State_Code"] = first_sf.get(
                    "Location_State_Code",
                    ""
                )

                data["GST_Order_Address_State"] = first_sf.get(
                    "GST_Order_Address_State",
                    ""
                )

                data["Buy_from_Vendor_No"] = first_sf.get(
                    "Buy_from_Vendor_No",
                    ""
                )

                data["Vendor_Invoice_No"] = first_sf.get(
                    "Vendor_Invoice_No",
                    ""
                )

                data["Currency_Code"] = first_sf.get(
                    "Currency_Code",
                    ""
                )

                data["Gen_Bus_Posting_Group"] = first_sf.get(
                    "Gen_Bus_Posting_Group",
                    ""
                )
        
            # =================================================
            # GST Vendor Type
            # =================================================
            if seller_gstin and len(seller_gstin) >= 2:

                data["GST_Vendor_Type"] = "Registered"

            else:

                data["GST_Vendor_Type"] = "Un-Registered"

            # =================================================
            # items[] ONLY FROM HSN API
            # =================================================
            for item in data.get("items", []):

                item_desc = clean(
                    item.get("Description")
                )

                hsn_info = hsn_map.get(
                    (order_no, item_desc)
                )

                if hsn_info:

                    item["PurchaseOrderNo"] = hsn_info.get(
                        "PurchaseOrderNo",
                        ""
                    )

                    item["PartSpecification"] = hsn_info.get(
                        "PartSpecification",
                        ""
                    )

                    item["ProductNo"] = (

                        hsn_info.get("ProductNo")

                        or hsn_info.get(
                            "Nav_Item_No",
                            ""
                        )
                    )

                    item["HSN_Type"] = hsn_info.get(
                        "HSN_Type",
                        ""
                    )

                    item["HSN_Percentage_Description"] = hsn_info.get(
                        "HSN_Percentage_Description",
                        ""
                    )
                    item["Nav_Item_No"] = hsn_info.get(
                        "Nav_Item_No",
                        ""
                    )

                    item["GST_%"] = hsn_info.get(
                        "TaxPercentage",
                        ""
                    )

                    item["TaxPercentage"] = hsn_info.get(
                        "TaxPercentage",
                        ""
                    )

                    # ============================================
                    # QUANTITY
                    # ============================================
                    qty_value = hsn_info.get(
                        "Quantity",
                        item.get("quantity", 0)
                    )

                    try:

                        qty_value = str(
                            qty_value
                        ).lower()

                        qty_value = qty_value.replace(
                            ",",
                            ""
                        )

                        match = re.search(
                            r"\d+(\.\d+)?",
                            qty_value
                        )

                        if match:

                            qty_value = float(
                                match.group()
                            )

                            if qty_value.is_integer():
                                qty_value = int(
                                    qty_value
                                )

                        else:

                            qty_value = 0

                    except:

                        qty_value = 0

                    item["Quantity"] = qty_value

                    # ============================================
                    # RATE
                    # ============================================
                    rate = item.get(
                        "rate",
                        0
                    )

                    try:

                        rate = str(rate).replace(
                            ",",
                            ""
                        )

                        rate = float(rate)

                    except:

                        rate = 0

                    # ============================================
                    # CALCULATE AMOUNT
                    # ============================================
                    try:

                        amount = qty_value * rate

                        if amount.is_integer():
                            amount = int(amount)

                        item["Amount"] = amount

                    except:

                        item["Amount"] = 0

                    # ============================================
                    # GST BASE AMOUNT
                    # ============================================
                    item["GST_Base_Amount"] = item.get(
                        "Amount",
                        0
                    )



                    # ============================================
                    # LINE NO FROM sl_no
                    # remove special chars and * 10000
                    # ============================================
                    sl_no = str(
                        item.get("sl_no", "")
                    )

                    sl_no = re.sub(
                        r"[^0-9]",
                        "",
                        sl_no
                    )

                    try:

                        sl_no_int = int(sl_no)

                        item["Line_No"] = sl_no_int # * 10000

                    except:

                        item["Line_No"] = 0
                        



                    # =====================================================
                    # COPY Line_No FROM items -> sf_items
                    # MATCH USING Nav_Item_No
                    # =====================================================
                    for sf in data.get("sf_items", []):

                        sf_nav_item = str(
                            sf.get("Nav_Item_No", "")
                        ).strip().lower()

                        if not sf_nav_item:
                            continue

                        matched_item = None

                        for item in data.get("items", []):

                            item_nav_item = str(
                                item.get("Nav_Item_No", "")
                            ).strip().lower()

                            if sf_nav_item == item_nav_item:

                                matched_item = item
                                break

                        if matched_item:

                            sf["Line_No"] = matched_item.get(
                                "Line_No",
                                0
                            )

                            print(
                                f"✅ Line_No COPIED: "
                                f"{sf_nav_item} -> "
                                f"{sf['Line_No']}"
                            )

                        else:

                            print(
                                f"❌ Line_No MATCH FAILED: "
                                f"{sf_nav_item}"
                            )



                    print(
                        f"✅ HSN UPDATED: "
                        f"{item.get('Description', '')}"
                    )

                else:

                    print(
                        f"❌ HSN NOT FOUND: "
                        f"{item.get('Description', '')}"
                    )
                    # add_error(
                    #     file_name=obj.get("file", ""),
                    #     entity=obj.get("entity", ""),
                    #     invoice_no=data.get("invoice_no", ""),
                    #     reason=f"Invoice part description not matching with SF part description/hsn not found "
                    # ) 



        print("====================================")
        print("✅ fetch_api_items COMPLETED")
        print("====================================")

        return grouped_data

    except Exception as e:

        print("❌ API ERROR:", e)

        return {}

# def fetch_api_items(order_numbers, json_data_list):
#     print("ENTERED fetch_api_items")
#     return {}
    
def apply_sf_item_defaults(api_rows):

    for row in api_rows:

        row.setdefault("Positive", "TRUE")
        row.setdefault("Suppressed_Action_Msg", 1)


        row.setdefault("Quantity_Base", 1)
        row.setdefault("Quantity_Invoiced_Base", 1)
        row.setdefault("Qty_per_Unit_of_Measure", 1)
        row.setdefault("Quantity", 1)

    return api_rows

def validate_json(obj, data, session_id=None):

    errors = []

    # =========================
    # Vendor Validation
    # =========================
    if not str(data.get("Nav_VendorCode") or "").strip(): 
        add_error( 
            session_id=session_id,
            file_name=obj.get("file", ""),
            entity=obj.get("entity", ""),
            invoice_no=data.get("invoice_no", ""),
            reason="Navision Vendor Code not mentioned in Vendor Master in Service First"
        )
        errors.append(
            "Navision Vendor Code not mentioned in Vendor Master in Service First"
        )

    goods_hsns = {
        str(x).strip()
        for x in data.get("hsn_type", {}).get("Goods", [])
    }

    # =========================
    # ITEMS VALIDATION
    # =========================
    for item in data.get("items", []):

        item_hsn = str(
            item.get("HSN")
            or item.get("hsn")
            or ""
        ).strip()

        if item_hsn not in goods_hsns:
            continue

        # if not str(item.get("ProductNo") or "").strip():
        #     errors.append(
        #         "ProductNo missing in Service First"
        #     )

        if not str(item.get("PartSpecification") or "").strip(): 
            # add_error(
            #     file_name=obj.get("file", ""),
            #     entity=obj.get("entity", ""),
            #     invoice_no=data.get("invoice_no", ""),
            #     reason= "Part Specification not mention in Service First at time of Part Receive"
            # )
            add_error(
                session_id=session_id,
                file_name=obj.get("file", ""),
                entity=obj.get("entity", ""),
                invoice_no=data.get("invoice_no", ""),
                reason="Part Specification not mention in Service First at time of Part Receive"
            )
            errors.append(
                "Part Specification not mention in Service First at time of Part Receive"
            )

        if (
            "Nav_Item_No" not in item
            or not str(item.get("Nav_Item_No") or "").strip()
            or str(item.get("Nav_Item_No")).lower() == "null"
        ):
            # add_error(
            #     file_name=obj.get("file", ""),
            #     entity=obj.get("entity", ""),
            #     invoice_no=data.get("invoice_no", ""),
            #     reason= "Navision Item No not mentioned in Part Master in Service First"
            # )
            add_error(
                session_id=session_id,
                file_name=obj.get("file", ""),
                entity=obj.get("entity", ""),
                invoice_no=data.get("invoice_no", ""),
                reason="Navision Item No not mentioned in Part Master in Service First"
            )
            errors.append(
                "Navision Item No not mentioned in Part Master in Service First"
            )

    # =========================
    # SF_ITEMS VALIDATION
    # =========================
    for sf_item in data.get("sf_items", []):

        if (
            "Nav_Item_No" not in sf_item
            or not str(sf_item.get("Nav_Item_No") or "").strip()
            or str(sf_item.get("Nav_Item_No")).lower() == "null"
        ):
            add_error(
                file_name=obj.get("file", ""),
                entity=obj.get("entity", ""),
                invoice_no=data.get("invoice_no", ""),
                reason=  "Navision Item No not mentioned in Part Master in Service First"
            )
            errors.append(
                "Navision Item No not mentioned in Part Master in Service First"
            )


        if (
            "Source_Ref._No" not in sf_item
            or not str(sf_item.get("Source_Ref._No") or "").strip()
            or str(sf_item.get("Source_Ref._No")).lower() == "null"
        ):
            add_error( 
                session_id=session_id,
                file_name=obj.get("file", ""),
                entity=obj.get("entity", ""),
                invoice_no=data.get("invoice_no", ""),
                reason=  "Part Specification not mention in Service First at time of Part Receive OR Navision Item No not mentioned in Part Master in Service First"
            )
            errors.append(
                "Part Specification not mention in Service First at time of Part Receive OR Navision Item No not mentioned in Part Master in Service First"
            )

    return list(set(errors))

def map_slno_from_items_to_sf_items(json_path):
    """
    Map sl_no from items -> sf_items using Nav_Item_No match
    and store it as Source_Ref._No
    """

    try:
        if not json_path or not os.path.exists(json_path):
            return

        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        items = data.get("items", [])
        sf_items = data.get("sf_items", [])

        if not items or not sf_items:
            return

        # Nav_Item_No -> sl_no mapping
        nav_map = {}

        for item in items:
            nav = str(item.get("Nav_Item_No", "")).strip().lower()
            sl_no = item.get("sl_no", "")

            if nav:
                nav_map[nav] = sl_no

        # Apply to sf_items
        for sf in sf_items:
            sf_nav = str(sf.get("Nav_Item_No", "")).strip().lower()

            if sf_nav in nav_map:
                sf["Source_Ref._No"] = nav_map[sf_nav]

        # save back JSON
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)

        print(f"✅ SL_NO mapped → {os.path.basename(json_path)}")

    except Exception as e:
        print(f"❌ SL_NO mapping error: {e}") 
def process_api_and_move_json(
    json_data_list=None,
    processed_json_folder=None,  
    settings=None,
    excel_folder=None,
    sf_pending_root=None,
    error_rows=None, 
    session_id=None
):
    print("process_api_and_move_json Started")
    if settings is None:
        settings = load_settings()

    # Default folders
    if processed_json_folder is None:
        processed_json_folder = os.path.join(
            settings["OUTPUT_FOLDER"],
            "Processed_Json"
        )

    if excel_folder is None:
        excel_folder = os.path.join(
            settings["OUTPUT_FOLDER"],
            "Excel"
        )

    os.makedirs(processed_json_folder, exist_ok=True)

    # =====================================================
    # LOAD JSON FILES IF NOT PASSED
    # =====================================================
    if json_data_list is None:

        json_folder = os.path.join(
            settings["OUTPUT_FOLDER"],
            "Json"
        )

        json_data_list = []

        if os.path.exists(json_folder):

            for root, _, files in os.walk(json_folder):

                for file in files:

                    if not file.lower().endswith(".json"):
                        continue

                    json_path = os.path.join(root, file)

                    try:
                        with open(json_path, "r", encoding="utf-8") as f:
                            data = json.load(f)

                        relative_path = os.path.relpath(
                            root,
                            json_folder
                        )

                        json_data_list.append({
                            "file": file,
                            "data": data,
                            "json_path": json_path,
                            "relative_path": relative_path,
                            "entity": relative_path.split(os.sep)[0] if relative_path != "." else "",
                            "doc_type": relative_path.split(os.sep)[1] if os.sep in relative_path else ""
                        })

                    except Exception as e:
                        print(f"❌ Failed to load {json_path}: {e}")

        # print(f"📄 Loaded {len(json_data_list)} JSON files.")
        
    # print("\n🚀 PROCESS API STARTED")

    os.makedirs(processed_json_folder, exist_ok=True)

    # if dao is None:
    #     print("❌ DAO IS NONE")
    #     return []

    # =====================================================
    # GET ORDER NUMBERS
    # =====================================================
    order_numbers = list(set([
        str(obj.get("data", {}).get("buyer_order_no", "")).strip()
        for obj in json_data_list
        if obj.get("data", {}).get("buyer_order_no", "")
    ]))

    # print("📦 ORDER NUMBERS:", order_numbers)

    # print("json_data_list")
    # print(order_numbers)
    # print(dao)


    if not order_numbers:
        return []

    # =====================================================
    # FETCH API DATA
    # =====================================================
    api_data_map = fetch_api_items(
        order_numbers=order_numbers,
        json_data_list=json_data_list
    )
    
    
    print("api_data_map")
    print(api_data_map)

    api_data_map = {
        str(k).strip().lower(): v
        for k, v in api_data_map.items()
    }



    processed_files = []

    # =====================================================
    # PATH CLEANER
    # =====================================================
    def clean_path(path):
        return os.path.normpath(str(path)).replace("/", "\\").strip("\\ ")

    # =====================================================
    # TEMPLATE PATH RESOLVER (FIXED - OPTION 2)
    # =====================================================
    def get_template_relative_path(obj):

        # =================================================
        # OPTION 2: FORCE SOURCE FROM JSON PATH
        # =================================================
        json_path = obj.get("json_path", "")
        source_folder = os.path.dirname(json_path) if json_path else ""

        pdf_root = settings.get("PDF_FOLDER", "")

        # print("pdf_root:", pdf_root)
        # print("source_folder:", source_folder)
        # print("json_path :",json_path)

        # =================================================
        # PRIORITY 1: DERIVE FROM PDF ROOT
        # =================================================
        if source_folder and pdf_root:

            try:
                src = clean_path(source_folder).replace("\\", "/")
                root = clean_path(pdf_root).replace("\\", "/")

                if root in src:
                    relative = src.split(root, 1)[-1].strip("/")
                    relative = relative.replace("/", "\\")

                    if relative:
                        # print(f"✅ FULL TEMPLATE PATH (FROM JSON PATH): {relative}")
                        return relative

            except Exception as e:
                print(f"⚠ PATH ERROR: {e}")

        # =================================================
        # PRIORITY 2
        # =================================================
        relative_path = obj.get("relative_path")
        # print("relative_path:", relative_path)

        if relative_path:
            relative_path = clean_path(relative_path)
            # print(f"✅ relative_path USED: {relative_path}")
            return relative_path

        # =================================================
        # PRIORITY 3
        # =================================================
        template_path = obj.get("template_path")
        # print("template_path:", template_path)

        if template_path:
            template_path = clean_path(template_path)
            # print(f"✅ template_path USED: {template_path}")
            return template_path

        # =================================================
        # PRIORITY 4
        # =================================================
        entity = str(obj.get("entity", "")).strip("\\/ ")
        doc_type = str(obj.get("doc_type", "")).strip("\\/ ")

        if entity and doc_type:
            final_path = clean_path(os.path.join(entity, doc_type))
            # print(f"✅ entity/doc_type USED: {final_path}")
            return final_path

        # =================================================
        # FALLBACK
        # =================================================
        fallback = clean_path(obj.get("entity", "Unknown"))
        # print(f"⚠ FALLBACK PATH USED: {fallback}")
        return fallback

    # =====================================================
    # LOOP JSON FILES
    # =====================================================
    for obj in json_data_list:

        try:
            data = obj["data"]

            order_no = str(data.get("buyer_order_no", "")).strip().lower()

            # print("\n====================================")
            # print(f"📄 PROCESSING: {obj['file']}")
            # print(f"📦 ORDER NO: {order_no}")
            # print("====================================")

            api_rows = api_data_map.get(order_no, [])
 
            # print(api_data_map)

            # if not api_rows:

            #     print(f"❌ API ROWS NOT FOUND: {order_no}")

            #     if error_rows is not None:

            #         error_rows.append({
            #             "FileName": os.path.splitext(
            #                 obj.get("file", "")
            #             )[0] + ".pdf",
            #             "Entity": obj.get("entity", ""),
            #             "InvoiceNo": data.get("invoice_no", ""),
            #             "Reason": "Part not yet received in Service First"
            #         })

            #     # ==========================================
            #     # MOVE PDF TO SF PENDING FOLDER
            #     # ==========================================
            #     try:

            #         processed_pdf = obj.get(
            #             "processed_pdf_path",
            #             ""
            #         )

            #         if (
            #             sf_pending_root
            #             and processed_pdf
            #             and os.path.exists(processed_pdf)
            #         ):

            #             pending_pdf = os.path.join(
            #                 sf_pending_root,
            #                 obj.get("entity", "DEFAULT"),
            #                 obj.get("doc_type", "DEFAULT"),
            #                 os.path.basename(processed_pdf)
            #             )

            #             os.makedirs(
            #                 os.path.dirname(pending_pdf),
            #                 exist_ok=True
            #             )

            #             shutil.move(
            #                 processed_pdf,
            #                 pending_pdf
            #             )

            #             print(
            #                 f"⚠ PDF MOVED TO SF PENDING → "
            #                 f"{pending_pdf}"
            #             )

            #     except Exception as move_error:

            #         print(
            #             f"❌ PENDING PDF MOVE FAILED: "
            #             f"{move_error}"
            #         )

            #     continue

            # APPLY DEFAULTS
            api_rows = apply_sf_item_defaults(api_rows)

            data["sf_items"] = api_rows

            first_sf = api_rows[0]

            data["Nav_VendorCode"] = first_sf.get("Nav_VendorCode", "")
            data["Pay_to_Vendor_No"] = first_sf.get("Nav_VendorCode", "")
            data["PaymentTermsName"] = first_sf.get("PaymentTermsName", "")


            # ==========================================
            # INLINE VALIDATION
            # ==========================================

            validation_failed = False

            # Vendor Validation
            if not str(data.get("Nav_VendorCode") or "").strip():

                add_error(
                    session_id=session_id,
                    file_name=obj.get("file", ""),
                    entity=obj.get("entity", ""),
                    invoice_no=data.get("invoice_no", ""),
                    reason="Navision Vendor Code not mentioned in Vendor Master in Service First"
                )

                validation_failed = True

            # Goods HSN List
            goods_hsns = {
                str(x).strip()
                for x in data.get("hsn_type", {}).get("Goods", [])
            }

            # Validate Goods Items
            for item in data.get("items", []):

                item_hsn = str(
                    item.get("HSN")
                    or item.get("hsn")
                    or ""
                ).strip()

                # Skip non-Goods items
                if item_hsn not in goods_hsns:
                    continue

                # Part Specification
                if not str(item.get("PartSpecification") or "").strip():

                    add_error(
                        session_id=session_id,
                        file_name=obj.get("file", ""),
                        entity=obj.get("entity", ""),
                        invoice_no=data.get("invoice_no", ""),
                        reason="Part Specification not mention in Service First at time of Part Receive"
                    )

                    validation_failed = True

            # Validate SF Items
            for sf_item in data.get("sf_items", []):

                nav_item = str(sf_item.get("Nav_Item_No") or "").strip()

                if not nav_item or nav_item.lower() == "null":

                    add_error(
                        session_id=session_id,
                        file_name=obj.get("file", ""),
                        entity=obj.get("entity", ""),
                        invoice_no=data.get("invoice_no", ""),
                        reason="Navision Item No not mentioned in Part Master in Service First"
                    )

                    validation_failed = True

            # Skip this file if validation failed
            if validation_failed:
                print(f"Validation failed : {obj['file']}")
                continue
            
            # =========================
            # PAYMENT TERMS + DUE DATE
            # =========================

            payment_term = str(
                data.get("payment_terms_code", "")
            ).strip() 
            
            payment_term = str(
                data.get("PaymentTermsName", "")
            ).strip()

            # Default
            if not payment_term:

                payment_term = "30 DAYS"

            data["payment_terms_code"] = payment_term

            try:

                invoice_date = str(
                    data.get("invoice_date", "")
                ).strip()

                match = re.search(
                    r"\d+",
                    payment_term
                )

                if match:

                    days = int(
                        match.group()
                    )

                    inv_dt = None

                    for fmt in (
                        "%d-%m-%Y",
                        "%d-%b-%y",
                        "%d-%b-%Y",
                        "%d/%m/%Y",
                        "%d.%m.%Y"
                    ):

                        try:

                            inv_dt = datetime.strptime(
                                invoice_date,
                                fmt
                            )

                            break

                        except:
                            pass

                    if inv_dt:

                        data["Due_Date"] = (
                            inv_dt +
                            timedelta(days=days)
                        ).strftime("%d-%m-%Y")

                        # print(
                        #     f"Payment Term : {payment_term}"
                        # )

                        # print(
                        #     f"Invoice Date : {invoice_date}"
                        # )

                        # print(
                        #     f"Due Date     : {data['Due_Date']}"
                        # )

            except Exception as e:

                print(
                    f"Due date calculation failed : {e}"
                )
            
            # =========================
            # PAYMENT TERMS + DUE DATE END
            # =========================
            data.setdefault("Status", "Open")
            data.setdefault(
                "Posting_Date",
                datetime.now().strftime("%Y-%m-%d")
            )
            
            # # CLEAN BUYER ORDER NO
            # buyer_order_no = str(
            #     data.get("buyer_order_no", "")
            # ).strip()

            # data["buyer_order_no"] = re.sub(
            #     r'^SPRPUR/',
            #     '',
            #     buyer_order_no,
            #     flags=re.IGNORECASE
            # )


            # ==========================================
            # VALIDATE JSON
            # ==========================================
            # validation_errors = validate_json(
            #     obj,
            #     data
            # )
  
            # if validation_errors:

            #     print(
            #         f"⚠ VALIDATION FAILED: "
            #         f"{obj['file']}"
            #     )

            #     print(
            #         "Errors:",
            #         validation_errors
            #     )

            #     continue


            # SAVE UPDATED JSON
            with open(obj["json_path"], "w", encoding="utf-8") as f:
                json.dump(data, f, indent=4, ensure_ascii=False)

            # =================================================
            # TEMPLATE PATH (FIXED)
            # =================================================
            relative_path = get_template_relative_path(obj)

            # print(f"📂 TEMPLATE PATH → {relative_path}")

            # CREATE DESTINATION FOLDER
            dest_dir = os.path.join(
                processed_json_folder,
                *relative_path.split("\\")
            )

            os.makedirs(dest_dir, exist_ok=True)

            dest_path = os.path.join(dest_dir, obj["file"])

            pdf_name = (
                os.path.splitext(obj["file"])[0] + ".pdf"
            ).lower()
 

            error_files = {
                str(err.get("FileName", "")).lower()
                for err in get_errors()
            }

            # print("ERROR FILES:")
            # print(error_files)

            if pdf_name in error_files:

                # print(
                #     f"⚠ ERROR FILE SKIPPED: {pdf_name}"
                # )

                continue

            pdf_name = (
                os.path.splitext(
                    obj["file"]
                )[0]
                + ".pdf"
            ).lower()

            error_files = {

                str(
                    err.get(
                        "FileName",
                        ""
                    )
                ).lower()

                for err in get_errors()
            }

            if pdf_name in error_files:

                # print(
                #     f"⚠ ERROR FILE SKIPPED: "
                #     f"{pdf_name}"
                # )

                continue

            shutil.move(
                obj["json_path"],
                dest_path
            )

            obj["processed_path"] = dest_path
            obj["data"] = data
            obj["relative_path"] = relative_path

            processed_files.append(obj)

            # print(f"✅ JSON MOVED → {dest_path}")

        except Exception as e:
            # print(f"❌ PROCESS FAILED: {obj.get('file')}")
            print("anbu")
            print(str(e))
            add_error( 
                session_id=session_id,
                file_name=os.path.splitext(
                    obj.get("file", "")
                )[0] + ".pdf",
                entity=obj.get("entity", ""),
                invoice_no=obj.get(
                    "data",
                    {}
                ).get(
                    "invoice_no",
                    ""
                ),
                reason="PO Not yet completed in Service First"
            )

            # print("ERRORS:")
            # for err in get_errors():
            #     print(err)

    # =====================================================
    # GENERATE EXCEL
    # =====================================================
    generate_excel(
        processed_json_folder=os.path.join(
            settings["OUTPUT_FOLDER"],
            "Processed_Json"
        ),
        excel_folder=os.path.join(
            settings["OUTPUT_FOLDER"],
            "Excel"
        ),
        generated_json_folder=os.path.join(
            settings["OUTPUT_FOLDER"],
            "Generated_Json"
        )
    )
 
    # =====================================================
    # MOVE COMPLETED PDFs
    # (Processed_Json + Generated_Json)
    # =====================================================
    pdf_root = settings.get("PDF_FOLDER", "").rstrip("\\/")

    completed_pdf_root = os.path.join(
        os.path.dirname(pdf_root),
        "Completed_PDFs"
    )

    processed_json_root = os.path.join(
        settings["OUTPUT_FOLDER"],
        "Processed_Json"
    )

    generated_json_root = os.path.join(
        settings["OUTPUT_FOLDER"],
        "Generated_Json"
    )

    os.makedirs(completed_pdf_root, exist_ok=True)

    # -----------------------------------------------------
    # Collect PDF names from both folders
    # -----------------------------------------------------
    pdf_names = set()

    for folder in [processed_json_root, generated_json_root]:

        if not os.path.exists(folder):
            continue

        for root, _, files in os.walk(folder):

            for file in files:

                if file.lower().endswith(".json"):
                    pdf_names.add(
                        os.path.splitext(file)[0] + ".pdf"
                    )

    print("=" * 80)
    print("TOTAL PDFs TO MOVE :", len(pdf_names))
    print("=" * 80)

    # -----------------------------------------------------
    # Move PDFs
    # -----------------------------------------------------
    for pdf_name in pdf_names:

        source_pdf = None

        # Search PDF recursively
        for root, _, files in os.walk(pdf_root):

            for file in files:

                if file.lower() == pdf_name.lower():

                    source_pdf = os.path.join(root, file)
                    break

            if source_pdf:
                break

        if not source_pdf:

            print(f"⚠ PDF NOT FOUND : {pdf_name}")
            continue

        relative_pdf = os.path.relpath(
            source_pdf,
            pdf_root
        )

        destination_pdf = os.path.join(
            completed_pdf_root,
            relative_pdf
        )

        os.makedirs(
            os.path.dirname(destination_pdf),
            exist_ok=True
        )

        # If destination exists, delete it first
        if os.path.exists(destination_pdf):
            print(f"Replacing existing PDF : {destination_pdf}")
            os.remove(destination_pdf)

        # Cut (move) the PDF from source to destination
        shutil.move(source_pdf, destination_pdf)

        print("=" * 80)
        print("PDF MOVED")
        print("FILE   :", pdf_name)
        print("SOURCE :", source_pdf)
        print("DEST   :", destination_pdf)
        print("=" * 80)

        print("=" * 80)
        print("PDF MOVED")
        print("FILE   :", pdf_name)
        print("SOURCE :", source_pdf)
        print("DEST   :", destination_pdf)
        print("=" * 80)
    return processed_files


# def generate_excel_old(
#     processed_json_folder,
#     excel_folder,
#     generated_json_folder 
# ):

#     import os
#     import json
#     import shutil

#     from pathlib import Path
#     from collections import defaultdict
#     from datetime import datetime
#     from openpyxl import Workbook

#     os.makedirs(excel_folder, exist_ok=True)
#     os.makedirs(generated_json_folder, exist_ok=True)

#     print("=" * 80)
#     print("GENERATE EXCEL STARTED")
#     print("=" * 80)

#     grouped = defaultdict(list)

#     # --------------------------------------------------
#     # Read JSON Files
#     # --------------------------------------------------
#     for json_file in Path(processed_json_folder).rglob("*.json"):

#         try:

#             with open(
#                 json_file,
#                 "r",
#                 encoding="utf-8"
#             ) as f:

#                 invoice = json.load(f)

#             invoice["json_path"] = str(json_file)

#             entity = invoice.get(
#                 "Entity",
#                 "Unknown"
#             )

#             template = invoice.get(
#                 "Template",
#                 "Default"
#             )

#             print(
#                 f"Loaded JSON : {json_file.name}"
#             )
#             print(
#                 f"Entity      : {entity}"
#             )
#             print(
#                 f"Template    : {template}"
#             )

#             grouped[
#                 (entity, template)
#             ].append(invoice)

#         except Exception as e:

#             print(
#                 f"Failed to read {json_file}"
#             )
#             print(e)

#     print(
#         f"\nGroups Found : {len(grouped)}"
#     )

#     generated_files = []

#     # --------------------------------------------------
#     # Sheet Source Mapping
#     # --------------------------------------------------
#     sheet_sources = {

#         "Purchase_Header":
#             lambda inv: [inv],

#         "Purchase_Line":
#             lambda inv: inv.get(
#                 "items",
#                 []
#             ),

#         "Reservation_Entry":
#             lambda inv: inv.get(
#                 "sf_items",
#                 []
#             )
#     }

#     # --------------------------------------------------
#     # Create Excel Per Group
#     # --------------------------------------------------
#     for (
#         entity,
#         template
#     ), invoices in grouped.items():

#         print("\n" + "=" * 80)
#         print(
#             f"PROCESSING : {entity}\\{template}"
#         )
#         print("=" * 80)

#         wb = Workbook()

#         wb.remove(
#             wb.active
#         )

#         # --------------------------------------------------
#         # Create Sheets
#         # --------------------------------------------------
#         for (
#             sheet_name,
#             headers
#         ) in settings[
#             "HEADER_ORDER"
#         ].items():

#             print(
#                 f"\nCreating Sheet : {sheet_name}"
#             )

#             ws = wb.create_sheet(
#                 sheet_name
#             )

#             # Header Row
#             ws.append(
#                 headers
#             )

#             json_map = settings[
#                 "JSON_MAP"
#             ].get(
#                 sheet_name,
#                 {}
#             )

#             print(
#                 f"Headers Count : {len(headers)}"
#             )

#             print(
#                 f"JSON MAP Count: {len(json_map)}"
#             )

#             source_func = (
#                 sheet_sources.get(
#                     sheet_name,
#                     lambda inv: [inv]
#                 )
#             )

#             # --------------------------------------------------
#             # Debug Mapping
#             # --------------------------------------------------
#             print("\nJSON MAP")

#             for (
#                 json_field,
#                 excel_column
#             ) in json_map.items():

#                 print(
#                     f"{json_field}"
#                     f" -> "
#                     f"{excel_column}"
#                 )

#             # --------------------------------------------------
#             # Write Data
#             # --------------------------------------------------
#             for invoice_index, invoice in enumerate(
#                 invoices,
#                 start=1
#             ):

#                 records = source_func(
#                     invoice
#                 )

#                 for record_index, record in enumerate(
#                     records,
#                     start=1
#                 ):

#                     row = []

#                     print(
#                         f"\nInvoice #{invoice_index}"
#                     )

#                     for header in headers:

#                         value = ""

#                         matched_field = None

#                         for (
#                             json_field,
#                             excel_column
#                         ) in json_map.items():

#                             if (
#                                 excel_column
#                                 == header
#                             ):

#                                 matched_field = (
#                                     json_field
#                                 )

#                                 value = record.get(
#                                     json_field,
#                                     ""
#                                 )

#                                 break

#                         row.append(
#                             value
#                         )

#                         # DEBUG
#                         if matched_field:

#                             print(
#                                 f"{header}"
#                                 f" <- "
#                                 f"{matched_field}"
#                                 f" = "
#                                 f"{value}"
#                             )
#                     print(row)
#                     ws.append(
#                         row
#                     )

#         # --------------------------------------------------
#         # Save Excel
#         # --------------------------------------------------
#         filename = (
#             f"{entity}_{template}_"
#             f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
#         )

#         excel_path = os.path.join(
#             excel_folder,
#             filename
#         )

#         wb.save(
#             excel_path
#         )

#         generated_files.append(
#             excel_path
#         )

#         print(
#             f"\nExcel Saved : {excel_path}"
#         )

#         # --------------------------------------------------
#         # Move JSON
#         # --------------------------------------------------
#         for invoice in invoices:

#             src = invoice.get(
#                 "json_path"
#             )

#             if (
#                 src
#                 and
#                 os.path.exists(src)
#             ):

#                 dst = os.path.join(
#                     generated_json_folder,
#                     os.path.basename(src)
#                 )

#                 shutil.move(
#                     src,
#                     dst
#                 )

#                 print(
#                     f"Moved JSON : {dst}"
#                 )

#     print("\n" + "=" * 80)
#     print(
#         f"TOTAL EXCEL FILES : {len(generated_files)}"
#     )
#     print("=" * 80)

#     return generated_files


from pathlib import Path
from collections import defaultdict
import json


def load_and_group_invoices(processed_json_folder):

    grouped = defaultdict(list)

    for json_file in Path(processed_json_folder).rglob("*.json"):

        try:

            with open(json_file, "r", encoding="utf-8") as f:
                invoice = json.load(f)

            invoice["json_path"] = str(json_file)

            entity = invoice.get("Entity", "Unknown")
            template = invoice.get("Template", "Default")

            grouped[(entity, template)].append(invoice)

        except Exception as e:

            # print(f"Failed to read {json_file}")
            print(e)

    return grouped
def get_cell_value(
    header,
    invoice,
    record,
    header_to_json,
    static_values
):

    # Static value
    if header in static_values:
        return static_values[header]

    json_field = header_to_json.get(header)

    if not json_field:
        return ""

    if isinstance(record, dict) and json_field in record:
        return record.get(json_field, "")

    return invoice.get(json_field, "")

# def populate_sheet(
#     ws,
#     sheet_name,
#     headers,
#     invoices,
#     entity,
#     template,
#     sheet_sources
# ):

#     ws.append(headers)

#     json_map = settings["JSON_MAP"].get(
#         sheet_name,
#         {}
#     )

#     header_to_json = {
#         excel_col: json_field
#         for json_field, excel_col in json_map.items()
#     }

#     # --------------------------------------------------
#     # Static Values Lookup
#     # --------------------------------------------------
#     static_key = f"{entity}\\{template}"

#     template_config = (
#         settings
#         .get("Static_Values", {})
#         .get(static_key, {})
#     )

#     static_values = template_config.get(
#         sheet_name,
#         {}
#     )

#     # --------------------------------------------------
#     # DEBUG
#     # --------------------------------------------------
#     print("\n" + "=" * 80)
#     print("ENTITY          :", repr(entity))
#     print("TEMPLATE        :", repr(template))
#     print("STATIC KEY      :", repr(static_key))
#     print("SHEET           :", sheet_name)

#     print(
#         "AVAILABLE KEYS  :",
#         list(
#             settings.get(
#                 "Static_Values",
#                 {}
#             ).keys()
#         )
#     )

#     print(
#         "CONFIG FOUND    :",
#         bool(template_config)
#     )

#     print(
#         "STATIC COUNT    :",
#         len(static_values)
#     )

#     print(
#         "STATIC VALUES   :",
#         static_values
#     )

#     print("=" * 80)

#     source_func = sheet_sources.get(
#         sheet_name,
#         lambda inv: [inv]
#     )

#     for invoice in invoices:

#         records = source_func(invoice)

#         if not records:
#             records = [{}]

#         for record in records:

#             row = []

#             for header in headers:

#                 value = get_cell_value(
#                     header=header,
#                     invoice=invoice,
#                     record=record,
#                     header_to_json=header_to_json,
#                     static_values=static_values
#                 )

#                 row.append(value)

#             ws.append(row)


def populate_sheet(
    ws,
    sheet_name,
    headers,
    invoices,
    entity,
    template,
    sheet_sources
):

    ws.append(headers)

    json_map = settings.get(
        "JSON_MAP",
        {}
    ).get(
        sheet_name,
        {}
    )

    header_to_json = {
        excel_col: json_field
        for json_field, excel_col in json_map.items()
    }

    # --------------------------------------------------
    # Static Values
    # --------------------------------------------------
    static_key = f"{entity}\\{template}"

    template_config = (
        settings
        .get("Static_Values", {})
        .get(static_key, {})
    )

    static_values = template_config.get(
        sheet_name,
        {}
    )

    # --------------------------------------------------
    # Debug
    # --------------------------------------------------
    print("\n" + "=" * 80)
    print("ENTITY       :", entity)
    print("TEMPLATE     :", template)
    print("STATIC KEY   :", static_key)
    print("SHEET        :", sheet_name)
    print("STATIC COUNT :", len(static_values))
    print("=" * 80)

    source_func = sheet_sources.get(
        sheet_name,
        lambda inv: [inv]
    )

    # --------------------------------------------------
    # Process Invoices
    # --------------------------------------------------
    for invoice_no, invoice in enumerate(
        invoices,
        start=1
    ):

        records = source_func(invoice)

        if not records:
            records = [{}]

        for record in records:

            row = []

            for header in headers:

                # ==========================================
                # Special Auto Number Fields
                # ==========================================
                if (
                    sheet_name == "Purchase_Header"
                    and header == "No."
                ):
                    value = invoice_no

                elif (
                    sheet_name == "Purchase_Line"
                    and header == "Document No."
                ):
                    value = invoice_no

                elif (
                    sheet_name == "Reservation_Entry"
                    and header == "Source ID"
                ):
                    value = invoice_no

                else:

                    value = get_cell_value(
                        header=header,
                        invoice=invoice,
                        record=record,
                        header_to_json=header_to_json,
                        static_values=static_values
                    )

                row.append(value)

            ws.append(row)

        # --------------------------------------------------
        # Debug
        # --------------------------------------------------
        print(
            f"Invoice #{invoice_no} processed "
            f"for {sheet_name}"
        )

    print(
        f"{sheet_name} rows written : "
        f"{ws.max_row - 1}"
    )
import os
import shutil


def move_generated_jsons(
    invoices,
    generated_json_folder
):

    os.makedirs(
        generated_json_folder,
        exist_ok=True
    )

    for invoice in invoices:

        src = invoice.get("json_path")

        if src and os.path.exists(src):

            dst = os.path.join(
                generated_json_folder,
                os.path.basename(src)
            )

            shutil.move(
                src,
                dst
            )
def generate_excel(
    processed_json_folder,
    excel_folder,
    generated_json_folder
):

    import os

    from datetime import datetime
    from openpyxl import Workbook

    os.makedirs(
        excel_folder,
        exist_ok=True
    )

    grouped = load_and_group_invoices(
        processed_json_folder
    )

    generated_files = []

    sheet_sources = {

        "Purchase_Header":
            lambda inv: [inv],

        "Purchase_Line":
            lambda inv: inv.get(
                "items",
                []
            ),

        "Reservation_Entry":
            lambda inv: inv.get(
                "sf_items",
                []
            )
    }

    print("=" * 80)
    print("GENERATE EXCEL STARTED")
    print("=" * 80)

    for (
        entity,
        template
    ), invoices in grouped.items():

        print(
            f"\nProcessing : "
            f"{entity} | {template}"
        )

        wb = Workbook()

        wb.remove(
            wb.active
        )

        # ----------------------------------
        # Create Sheets
        # ----------------------------------

        for (
            sheet_name,
            headers
        ) in settings[
            "HEADER_ORDER"
        ].items():

            excel_sheet_name = (
                sheet_name
                .replace("_", " ")
            )

            ws = wb.create_sheet(
                excel_sheet_name
            )

            populate_sheet(
                ws=ws,
                sheet_name=sheet_name,
                headers=headers,
                invoices=invoices,
                entity=entity,
                template=template,
                sheet_sources=sheet_sources
            )

        filename = (
            f"{entity}_"
            f"{template.replace('/', '_').replace(chr(92), '_')}_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            f".xlsx"
        )

        excel_path = os.path.join(
            excel_folder,
            filename
        )

        wb.save(
            excel_path
        )

        generated_files.append(
            excel_path
        )

        print(
            f"Saved : {excel_path}"
        )

        move_generated_jsons(
            invoices,
            generated_json_folder
        )

    print("\n" + "=" * 80)
    print(
        f"Total Excel Files : "
        f"{len(generated_files)}"
    )
    print("=" * 80)

    return generated_files